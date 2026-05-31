#!/usr/bin/env python3
"""
库存金额跟踪系统 (Inventory Tracking System)
============================================
功能：
1. 从上月底Excel文件读取期初库存余额（支持 .xlsx / .xlsb）
2. 连接SQL Server查询当月MTD物料事务（使用 pymssql，无需 ODBC 驱动）
3. 按TransType+RefType分类：Received / Consumed / Other Transaction
4. 按Item汇总Qty和AMT（金额直接使用 TotalPosted）
5. 计算期末余额 = 期初 + Received - Consumed - Other
6. 导出Excel报表（Project Summary + Summary + Detail 三页，Project Summary 在前）
7. 支持多站点（310/330/410）批量运行
8. 通过 SMTP 发送邮件（支持内网 Relay，无需 Outlook）
9. 通过 .env 外部化配置（DB、SMTP、邮件收件人、Infor API）

作者: WorkBuddy for NAI Group
日期: 2026-05-25 | 更新: 2026-05-31 (移除 config.ini，统一用 .env 配置)
"""

import os
import sys
import time
import glob
import warnings
import argparse
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime, timedelta
from dataclasses import dataclass
from typing import Dict, List, Optional, Tuple
from pathlib import Path

import json
import urllib.request
import urllib.error
import urllib.parse

import pymssql
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────
# TransType + RefType → Category 映射规则
# ──────────────────────────────────────────────────────────────
CATEGORY_MAP: Dict[Tuple[str, str], str] = {
    # Received (入库)
    ("R", "P"): "Received",   # PO Receipt
    ("W", "P"): "Received",   # PO Withdraw

    # Consumed (消耗)
    ("I", "J"): "Consumed",   # Job Issue / WIP Change
    ("W", "J"): "Consumed",   # Job Withdrawal / Return
    ("S", "O"): "Consumed",   # Order Ship

    # Other Transaction (其他)
    ("A", "I"): "Other",      # Adjustment
    ("M", "I"): "Other",      # Stock Move
    ("G", "I"): "Other",      # Misc Receipt
    ("H", "I"): "Other",      # Misc Issue
    ("C", "J"): "Other",      # Job Complete
    ("F", "J"): "Other",      # Job Finish
    ("N", "J"): "Other",      # Job Labor / Next Operation
    ("W", "R"): "Other",      # RMA Withdraw
}

TRANS_DESCRIPTIONS: Dict[Tuple[str, str], str] = {
    ("R", "P"): "PO Receipt",
    ("W", "P"): "PO Withdraw",
    ("I", "J"): "Job Issue / WIP Change",
    ("W", "J"): "Job Withdrawal / Return",
    ("S", "O"): "Order Ship",
    ("A", "I"): "Adjustment",
    ("M", "I"): "Stock Move",
    ("G", "I"): "Misc Receipt",
    ("H", "I"): "Misc Issue",
    ("C", "J"): "Job Complete",
    ("F", "J"): "Job Finish",
    ("N", "J"): "Job Labor / Next Operation",
    ("W", "R"): "RMA Withdraw",
}

SITE_NAMES = {"310": "Plant1", "330": "Plant2", "410": "PNG"}
# Site 330 的原始货币是 CNY，需要按汇率转为 USD
SITE_CURRENCY = {"310": ("USD", 1.0), "330": ("CNY->USD", 1 / 6.838784), "410": ("USD", 1.0)}


@dataclass
class ItemBalance:
    item: str
    description: str = ""
    project_code: str = ""
    prev_balance: float = 0.0
    prev_qty: float = 0.0
    recv_qty: float = 0.0
    recv_amt: float = 0.0
    cons_qty: float = 0.0
    cons_amt: float = 0.0
    other_qty: float = 0.0
    other_amt: float = 0.0

    @property
    def balance(self) -> float:
        return self.prev_balance + self.recv_amt + self.cons_amt + self.other_amt

    @property
    def net_qty(self) -> float:
        return self.prev_qty + self.recv_qty + self.cons_qty + self.other_qty


# ──────────────────────────────────────────────────────────────
# 样式常量
# ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
BALANCE_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


class InventoryTracker:
    def __init__(
        self,
        server: str = r"SUZVPRINT01\CUSTOMSSYS",
        database: str = "csi_datawarehouse",
        username: str = "sa",
        password: str = "xxVcDW9ED24YWX",
        port: int = 1433,
        site_ref: str = "310",
        prev_balance_file: Optional[str] = None,
        output_file: Optional[str] = None,
    ):
        self.server = server
        self.database = database
        self.username = username
        self.password = password
        self.port = port
        self.site_ref = site_ref
        self.prev_balance_file = prev_balance_file

        # Currency: site 330 is CNY, convert to USD
        curr_info = SITE_CURRENCY.get(site_ref, ("USD", 1.0))
        self.currency_label = curr_info[0]
        self.fx_rate = curr_info[1]
        self.usd_symbol = "$"

        site_label = SITE_NAMES.get(site_ref, site_ref)
        today = datetime.today()
        self.period_str = today.strftime("%Y-%m")
        default_name = f"Inventory_Balance_{site_label}_{today.strftime('%Y%m')}.xlsx"
        self.output_file = output_file or default_name

    def _get_connection(self):
        """使用 pymssql 建立数据库连接（无需 ODBC 驱动）"""
        # SQL Server 命名实例格式：host\instance，pymssql 需拆分
        host = self.server
        instance = None
        if "\\" in host:
            host, instance = host.split("\\", 1)

        kwargs = dict(
            server=host,
            user=self.username,
            password=self.password,
            database=self.database,
            port=self.port,
            tds_version="7.4",
            login_timeout=15,
            conn_properties="SET TEXTSIZE 65536",
        )
        if instance:
            # pymssql 通过 server="host\\instance" 格式支持命名实例
            kwargs["server"] = f"{host}\\{instance}"
            del kwargs["port"]  # 命名实例时不指定端口，由 SQL Browser 解析

        return pymssql.connect(**kwargs)

    # ──────────────────────────────────────────────────────────
    # 1. 读取期初库存（Excel）
    # ──────────────────────────────────────────────────────────
    def load_previous_balance(self) -> pd.DataFrame:
        """
        从库存Excel读取上月末余额。
        支持两种列名格式：
          A) 标准格式：Item, Prev_Qty, Prev_Balance
          B) Infor导出格式：Item, Per, Unitcost（单价）→ Prev_Balance = Per × Unitcost
          C) 旧Infor格式：Item, Per, Unitscost（扩展成本）→ 直接使用
        """
        if not self.prev_balance_file or not os.path.exists(self.prev_balance_file):
            print(f"  ⚠️  期初余额文件未找到：{self.prev_balance_file}")
            print("     期初余额将设为 0")
            return pd.DataFrame(columns=["Item", "Prev_Qty", "Prev_Balance", "Description"])

        filepath = self.prev_balance_file
        # 自动检测是否需要跳过第一行（site 410 等文件第一行是标题）
        skip = 0
        if filepath.lower().endswith(".xlsb"):
            skip = 1
        else:
            # 读取前2行，检查列名是否全为 "Unnamed"（说明第一行是标题行，不是列名）
            test_df = pd.read_excel(filepath, nrows=2)
            if all(c.startswith("Unnamed") for c in test_df.columns):
                skip = 1

        df = pd.read_excel(filepath, skiprows=skip)
        df.columns = [c.strip() for c in df.columns]

        # 智能识别列名并计算期初金额
        if "Prev_Qty" in df.columns and "Prev_Balance" in df.columns:
            # 标准格式
            qty_col, amt_col = "Prev_Qty", "Prev_Balance"
            calc_extended = False
        elif "Per" in df.columns and "Unitscost" in df.columns:
            # Infor API 格式：Unitscost 是扩展金额（= Units × Unitcost，由 API 直接提供）
            # Prev_Balance 直接使用 Unitscost，无需计算
            qty_col, amt_col = "Per", "Unitscost"
            calc_extended = False
        elif "Per" in df.columns and "Unitcost" in df.columns:
            # 新Infor格式：Unitcost 是单价，需要 Per × Unitcost
            qty_col = "Per"
            calc_extended = True
            print("  ℹ️  检测到 Unitcost（单价），将计算 Per × Unitcost 作为期初金额")
        else:
            raise ValueError(
                f"无法识别期初余额列名。现有列：{list(df.columns)}\n"
                f"需要包含 (Item + Prev_Qty + Prev_Balance) 或 (Item + Per + Unitcost)"
            )

        df["Item"] = df["Item"].astype(str).str.strip().str.upper()
        df = df[df["Item"] != "NAN"].copy()
        df["Prev_Qty"] = pd.to_numeric(df[qty_col], errors="coerce").fillna(0)

        if calc_extended:
            unit_cost = pd.to_numeric(df["Unitcost"], errors="coerce").fillna(0)
            df["Prev_Balance"] = df["Prev_Qty"] * unit_cost
        else:
            df["Prev_Balance"] = pd.to_numeric(df[amt_col], errors="coerce").fillna(0)

        # 尝试保留 Description
        desc_col = None
        for c in ["Description", "ue_GDL_Description"]:
            if c in df.columns:
                desc_col = c
                break
        if desc_col:
            df["Description"] = df[desc_col].astype(str).str.strip()
        else:
            df["Description"] = ""

        df = df[["Item", "Prev_Qty", "Prev_Balance", "Description"]].copy()

        # Currency conversion for non-USD sites
        if self.fx_rate != 1.0:
            df["Prev_Balance"] = df["Prev_Balance"] * self.fx_rate
            print(f"  💱 汇率转换 {self.currency_label} ÷ 6.838784 = USD")

        total_bal = df["Prev_Balance"].sum()
        print(f"  ✅ 已加载 {len(df)} 个Item，期初总金额: ${total_bal:,.2f}")
        return df

    # ──────────────────────────────────────────────────────────
    # 2. 查询数据库（MTD事务）
    # ──────────────────────────────────────────────────────────
    def fetch_mtd_transactions(self) -> pd.DataFrame:
        today = datetime.today()
        month_start = today.replace(day=1)

        query = f"""
        SELECT
            m.[SiteRef], m.[TransNum], m.[TransDate], m.[TransType], m.[RefType], m.[Backflush],
            m.[Whse], m.[Loc], m.[Lot], m.[Wc], m.[RefNum], m.[RefLineSuf], m.[RefRelease],
            m.[Item], m.[ue_GDL_Description],
            m.[Qty], m.[TotalPosted],
            m.[RowPointer], m.[RecordDate], m.[DocumentNumber],
            dbo.GET_Customer_Formate_ProjectCode(
                CAST(m.[SiteRef] AS int), m.[Item]
            ) AS ProjectCode
        FROM [csi_datawarehouse].[dbo].[SLMatltrans] m
        INNER JOIN [csi_datawarehouse].[dbo].[SLItems] i
            ON m.[Item] = i.[Item] AND m.[SiteRef] = i.[SiteRef]
        WHERE m.[SiteRef] = '{self.site_ref}'
          AND m.[TransDate] >= '{month_start.strftime('%Y-%m-%d')}'
          AND i.[PMTCode] = 'P'
        ORDER BY m.[TransDate], m.[TransNum]
        """

        print(f"  📊 查询 {self.database} Site {self.site_ref} 采购物料 (PMTCode=P) ({month_start.strftime('%Y-%m-%d')} ~)")

        try:
            conn = self._get_connection()
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                df = pd.read_sql(query, conn)
            conn.close()
        except Exception as e:
            print(f"  ❌ 数据库连接失败：{e}")
            sys.exit(1)

        df["Item"] = df["Item"].astype(str).str.strip().str.upper()
        df = df[df["Item"] != "NAN"].copy()
        df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0)
        df["TotalAmt"] = pd.to_numeric(df["TotalPosted"], errors="coerce").fillna(0)

        # Currency conversion for non-USD sites
        if self.fx_rate != 1.0:
            df["TotalAmt"] = df["TotalAmt"] * self.fx_rate

        print(f"  ✅ 查询到 {len(df):,} 条事务记录，{df['Item'].nunique():,} 个Item")
        return df

    # ──────────────────────────────────────────────────────────
    # 3. 分类
    # ──────────────────────────────────────────────────────────
    def classify(self, row: pd.Series) -> str:
        key = (str(row.get("TransType", "")).strip().upper(),
               str(row.get("RefType", "")).strip().upper())
        return CATEGORY_MAP.get(key, "Other")

    def _desc(self, tt: str, rt: str) -> str:
        return TRANS_DESCRIPTIONS.get((str(tt).strip().upper(), str(rt).strip().upper()),
                                      f"{tt}/{rt}")

    # ──────────────────────────────────────────────────────────
    # 4. 按 Item 汇总
    # ──────────────────────────────────────────────────────────
    def calculate_balances(self, prev_df: pd.DataFrame, trans_df: pd.DataFrame):
        trans_df["Category"] = trans_df.apply(self.classify, axis=1)
        trans_df["TransDesc"] = trans_df.apply(
            lambda r: self._desc(r["TransType"], r["RefType"]), axis=1
        )

        # Detail
        detail_cols = [
            "Item", "TransDate", "TransNum", "TransType", "RefType",
            "Category", "TransDesc", "Qty", "TotalAmt",
            "Whse", "Loc", "Lot", "RefNum", "RefLineSuf",
        ]
        available = [c for c in detail_cols if c in trans_df.columns]
        detail_df = trans_df[available].copy()
        detail_df["TransDate"] = pd.to_datetime(detail_df["TransDate"]).dt.strftime("%Y-%m-%d")

        # Build description + ProjectCode lookup from prev_df
        desc_map = {}
        if "Description" in prev_df.columns:
            for _, row in prev_df.iterrows():
                desc_map[str(row["Item"])] = row["Description"]

        # Build ProjectCode lookup from transaction data (via GET_Customer_Formate_ProjectCode)
        project_code_map: Dict[str, str] = {}
        if "ProjectCode" in trans_df.columns:
            for _, row in trans_df.iterrows():
                pc = str(row.get("ProjectCode", "")).strip()
                if pc and pc != "NAN" and pc != "NONE":
                    project_code_map[str(row["Item"])] = pc

        # Summary aggregation
        items: Dict[str, ItemBalance] = {}
        for _, row in prev_df.iterrows():
            item = str(row["Item"])
            items[item] = ItemBalance(
                item=item,
                prev_balance=row["Prev_Balance"],
                prev_qty=row["Prev_Qty"],
                description=desc_map.get(item, ""),
                project_code=project_code_map.get(item, ""),
            )

        for _, row in trans_df.iterrows():
            item = str(row["Item"])
            cat = row["Category"]
            qty, amt = row["Qty"], row["TotalAmt"]
            if item not in items:
                items[item] = ItemBalance(item=item, project_code=project_code_map.get(item, ""))
            ib = items[item]
            if cat == "Received":
                ib.recv_qty += qty; ib.recv_amt += amt
            elif cat == "Consumed":
                ib.cons_qty += qty; ib.cons_amt += amt
            else:
                ib.other_qty += qty; ib.other_amt += amt

        summary_data = []
        for item, ib in items.items():
            summary_data.append({
                "ProjectCode": ib.project_code,
                "Item": item,
                "Description": ib.description,
                "Prev_Qty": round(ib.prev_qty, 4),
                "Prev_Balance": round(ib.prev_balance, 2),
                "Received_Qty": round(ib.recv_qty, 4),
                "Received_AMT": round(ib.recv_amt, 2),
                "Consumed_Qty": round(ib.cons_qty, 4),
                "Consumed_AMT": round(ib.cons_amt, 2),
                "Other_Qty": round(ib.other_qty, 4),
                "Other_AMT": round(ib.other_amt, 2),
                "Balance_Qty": round(ib.net_qty, 4),
                "Balance_AMT": round(ib.balance, 2),
            })

        summary_df = pd.DataFrame(summary_data)
        # 按期末余额从大到小排序
        summary_df = summary_df.sort_values("Balance_AMT", ascending=False).reset_index(drop=True)
        return summary_df, detail_df

    # ──────────────────────────────────────────────────────────
    # 5. 导出 Excel
    # ──────────────────────────────────────────────────────────
    def export_excel(self, summary_df: pd.DataFrame, detail_df: pd.DataFrame) -> None:
        output_path = Path(self.output_file)
        today = datetime.today()
        site_label = SITE_NAMES.get(self.site_ref, self.site_ref)

        # ── 先用 pandas 写入原始数据（快速，无样式）──
        print(f"  📝 导出 ({len(summary_df):,} summary + {len(detail_df):,} detail rows)...")

        # 创建空白 workbook 并手动写入，确保完全控制行位置
        wb = Workbook()

        # ═══════════════════════════════════════════════════════════
        # Sheet 1: Project Code Summary (按项目代码汇总) ← 放第一个
        # ═══════════════════════════════════════════════════════════
        ws_proj = wb.active
        ws_proj.title = "Project Summary"

        # 按 ProjectCode 分组汇总（不含 Items 计数）
        proj_group = summary_df.groupby("ProjectCode", dropna=False).agg(
            Prev_Balance=("Prev_Balance", "sum"),
            Received_AMT=("Received_AMT", "sum"),
            Consumed_AMT=("Consumed_AMT", "sum"),
            Other_AMT=("Other_AMT", "sum"),
            Balance_AMT=("Balance_AMT", "sum"),
        ).reset_index()
        proj_group = proj_group.sort_values("Balance_AMT", ascending=False).reset_index(drop=True)
        proj_group["ProjectCode"] = proj_group["ProjectCode"].fillna("")

        # Row 1: 标题
        ws_proj.merge_cells("A1:H1")
        ws_proj["A1"] = f"按项目代码汇总 - Site {self.site_ref} ({site_label})"
        ws_proj["A1"].font = TITLE_FONT
        ws_proj["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_proj.row_dimensions[1].height = 28

        # Row 2: 表头
        proj_headers = ["Project Code", "4/30 Balance",
                        "MTD Received", "MTD Consumption", "MTD Other Transaction",
                        "MTD Daily Balance", "公式", ""]
        for col_idx, col_name in enumerate(proj_headers, 1):
            c = ws_proj.cell(row=2, column=col_idx, value=col_name)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER
        ws_proj.merge_cells("G2:H2")
        ws_proj.row_dimensions[2].height = 25

        # Row 3+: 数据行
        for row_data in proj_group.itertuples(index=False):
            ws_proj.append(list(row_data))

        # 为每行添加公式说明
        data_start = 3
        data_end = ws_proj.max_row
        for r in range(data_start, data_end + 1):
            ws_proj.cell(row=r, column=7, value=f"=B{r}+C{r}+D{r}+E{r}")
            ws_proj.cell(row=r, column=7).number_format = "#,##0.00"
            ws_proj.cell(row=r, column=8, value="4/30+Recv+Cons+Other")
            ws_proj.cell(row=r, column=8).font = Font(size=9, color="666666")
            for col_idx in [2, 3, 4, 5, 6]:
                ws_proj.cell(row=r, column=col_idx).number_format = "#,##0.00"

        # Total row
        total_r = data_end + 1
        ws_proj.cell(row=total_r, column=1, value="TOTAL")
        ws_proj.cell(row=total_r, column=1).font = Font(bold=True)
        ws_proj.cell(row=total_r, column=1).fill = TOTAL_FILL
        ws_proj.cell(row=total_r, column=1).border = THIN_BORDER
        proj_totals = {
            2: proj_group["Prev_Balance"].sum(),
            3: proj_group["Received_AMT"].sum(),
            4: proj_group["Consumed_AMT"].sum(),
            5: proj_group["Other_AMT"].sum(),
            6: proj_group["Balance_AMT"].sum(),
        }
        for col_idx, total_val in proj_totals.items():
            c = ws_proj.cell(row=total_r, column=col_idx)
            c.value = round(total_val, 2)
            c.font = Font(bold=True)
            c.fill = TOTAL_FILL
            c.border = THIN_BORDER
            c.number_format = "#,##0.00"
        t_prev = proj_totals[2]; t_recv = proj_totals[3]; t_cons = proj_totals[4]; t_other = proj_totals[5]
        ws_proj.cell(row=total_r, column=7, value=round(t_prev + t_recv - t_cons - t_other, 2))
        ws_proj.cell(row=total_r, column=7).font = Font(bold=True)
        ws_proj.cell(row=total_r, column=7).fill = TOTAL_FILL
        ws_proj.cell(row=total_r, column=7).number_format = "#,##0.00"

        proj_widths = [18, 16, 16, 16, 14, 16, 20, 20]
        for i, w in enumerate(proj_widths, 1):
            ws_proj.column_dimensions[get_column_letter(i)].width = w
        ws_proj.freeze_panes = "A3"

        # ═══════════════════════════════════════════════════════════
        # Sheet 2: Summary (Item 明细)
        # ═══════════════════════════════════════════════════════════
        ws = wb.create_sheet("Summary")

        ws.merge_cells("A1:M1")
        ws["A1"] = f"库存金额跟踪报表 - Site {self.site_ref} ({site_label}) - 采购物料 (PMTCode=P)"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

        ws.merge_cells("A2:M2")
        ws["A2"] = (f"报告周期：{today.strftime('%Y年%m月')} 1日 - {today.strftime('%m月%d日')}   |   "
                     f"数据截至：{today.strftime('%Y-%m-%d %H:%M')}   |   "
                     f"仅含采购物料")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws["A2"].font = Font(size=10, italic=True, color="666666")

        headers = ["Project Code", "Item", "Description", "4/30 Qty", "4/30 Balance",
                    "MTD Received Qty", "MTD Received", "MTD Consumption Qty", "MTD Consumption",
                    "MTD Other Qty", "MTD Other Transaction", "MTD Daily Balance Qty", "MTD Daily Balance"]
        ws.append(headers)
        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=3, column=col_num)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER

        for row_data in summary_df.itertuples(index=False):
            ws.append(list(row_data))

        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        ws.cell(row=total_row, column=1).fill = TOTAL_FILL
        ws.cell(row=total_row, column=1).border = THIN_BORDER
        col_to_field = {
            4:  ("Prev_Qty",     "#,##0.0000"),
            5:  ("Prev_Balance", "#,##0.00"),
            6:  ("Received_Qty", "#,##0.0000"),
            7:  ("Received_AMT", "#,##0.00"),
            8:  ("Consumed_Qty", "#,##0.0000"),
            9:  ("Consumed_AMT", "#,##0.00"),
            10: ("Other_Qty",    "#,##0.0000"),
            11: ("Other_AMT",    "#,##0.00"),
            12: ("Balance_Qty",  "#,##0.0000"),
            13: ("Balance_AMT",  "#,##0.00"),
        }
        for col_idx, (field, fmt) in col_to_field.items():
            total_val = summary_df[field].sum() if field in summary_df.columns else 0
            c = ws.cell(row=total_row, column=col_idx)
            c.value = round(total_val, 4 if "Qty" in field else 2)
            c.font = Font(bold=True)
            c.fill = TOTAL_FILL
            c.border = THIN_BORDER
            c.number_format = fmt

        widths = [16, 20, 40, 12, 15, 14, 15, 14, 15, 12, 14, 12, 15]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = "A4"
        ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{total_row - 1}"
        ws.row_dimensions[1].height = 28
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 30

        # ═══════════════════════════════════════════════════════════
        # Sheet 3: Detail
        # ═══════════════════════════════════════════════════════════
        ws_det = wb.create_sheet("Detail")

        det_cols = len(detail_df.columns)
        ws_det.merge_cells(f"A1:{get_column_letter(det_cols)}1")
        ws_det["A1"] = f"物料事务明细 - Site {self.site_ref} ({site_label}) - 采购物料 (PMTCode=P)"
        ws_det["A1"].font = TITLE_FONT
        ws_det["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_det.row_dimensions[1].height = 28

        for col_idx, col_name in enumerate(detail_df.columns, 1):
            c = ws_det.cell(row=2, column=col_idx, value=col_name)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER
        ws_det.row_dimensions[2].height = 25

        for row_data in detail_df.itertuples(index=False):
            ws_det.append(list(row_data))

        ws_det.freeze_panes = "A3"
        ws_det.auto_filter.ref = f"A2:{get_column_letter(det_cols)}{ws_det.max_row}"

        for col_idx in range(1, det_cols + 1):
            col_letter = get_column_letter(col_idx)
            col_name = detail_df.columns[col_idx - 1]
            if col_name in ("Item", "TransNum", "RefNum"):
                ws_det.column_dimensions[col_letter].width = 18
            elif col_name in ("TransDate", "TransType", "RefType", "Category", "TransDesc", "Whse", "Loc", "Lot"):
                ws_det.column_dimensions[col_letter].width = 14
            elif col_name == "Description":
                ws_det.column_dimensions[col_letter].width = 40
            elif col_name in ("Qty", "TotalAmt"):
                ws_det.column_dimensions[col_letter].width = 14
            else:
                ws_det.column_dimensions[col_letter].width = 12

        wb.save(output_path)
        print(f"  📄 报表已导出：{output_path.absolute()}")
        print(f"     Summary: {len(summary_df):,} Items (PMTCode=P)")
        print(f"     Detail: {len(detail_df):,} records")
        print(f"     Project Summary: {len(proj_group):,} Project Codes")

    # ──────────────────────────────────────────────────────────
    # 主流程
    # ──────────────────────────────────────────────────────────
    def run(self):
        site_label = SITE_NAMES.get(self.site_ref, self.site_ref)
        print(f"\n{'='*60}")
        print(f"  库存跟踪 - Site {self.site_ref} ({site_label})")
        print(f"{'='*60}")

        print("\n📁 Step 1: 加载期初库存余额...")
        prev_df = self.load_previous_balance()

        print("\n🗄️  Step 2: 查询当月MTD物料事务...")
        trans_df = self.fetch_mtd_transactions()

        print("\n📊 Step 3: 分类汇总计算...")
        summary_df, detail_df = self.calculate_balances(prev_df, trans_df)

        print("\n📤 Step 4: 导出Excel报表...")
        self.export_excel(summary_df, detail_df)

        # Quick summary
        prev_total = summary_df["Prev_Balance"].sum()
        recv_total = summary_df["Received_AMT"].sum()
        cons_total = summary_df["Consumed_AMT"].sum()
        other_total = summary_df["Other_AMT"].sum()
        bal_total = prev_total + recv_total + cons_total + other_total

        print(f"\n  ┌─────────────────────────────────────────┐")
        print(f"  │  期初余额:   ${prev_total:>15,.2f}       │")
        print(f"  │  + Received: ${recv_total:>15,.2f}       │")
        print(f"  │  + Consumed: ${cons_total:>15,.2f}       │")
        print(f"  │  + Other:    ${other_total:>15,.2f}       │")
        print(f"  │  = Balance:  ${bal_total:>15,.2f}       │")
        print(f"  └─────────────────────────────────────────┘")
        print(f"\n✅ Site {self.site_ref} ({site_label}) 完成！")
        return summary_df, detail_df


# ──────────────────────────────────────────────────────────────
# 多站点批量运行
# ──────────────────────────────────────────────────────────────
def run_all_sites(
    server: str, database: str, username: str, password: str, port: int = 1433,
    prev_dir: str = "Previous Balance", output_dir: Optional[str] = None,
    sites: Optional[List[str]] = None,
):
    """批量运行所有站点"""
    if sites is None:
        sites = ["310", "330", "410"]

    # 自动匹配期初文件到站点（优先匹配新命名 site XXX.xlsx）
    prev_files = {"310": None, "330": None, "410": None}
    prev_path = Path(prev_dir)
    # 第一轮：精确匹配 "site XXX.xlsx"
    for site in ["310", "330", "410"]:
        exact = prev_path / f"site {site}.xlsx"
        if exact.exists():
            prev_files[site] = str(exact)
    # 第二轮：模糊匹配旧命名 (Plant1/Plant2/PNG)
    if not all(prev_files.values()):
        for f in prev_path.iterdir():
            fname = f.name.upper()
            if not fname.endswith((".XLSX", ".XLSB")):
                continue
            if "PLANT1" in fname and not prev_files["310"]:
                prev_files["310"] = str(f)
            elif "PLANT2" in fname and not prev_files["330"]:
                prev_files["330"] = str(f)
            elif "PNG" in fname and not prev_files["410"]:
                prev_files["410"] = str(f)

    out_dir = Path(output_dir) if output_dir else Path(".")
    out_dir.mkdir(parents=True, exist_ok=True)

    all_summaries = {}
    for site in sites:
        pf = prev_files.get(site)
        if not pf or not os.path.exists(pf):
            print(f"\n⚠️  Site {site}: 未找到期初文件，跳过")
            continue

        tracker = InventoryTracker(
            server=server, database=database,
            username=username, password=password, port=port,
            site_ref=site, prev_balance_file=pf,
            output_file=str(out_dir / f"Inventory_Balance_{SITE_NAMES.get(site, site)}_{datetime.today().strftime('%Y%m')}.xlsx"),
        )
        summary_df, detail_df = tracker.run()
        summary_df["Site"] = site
        all_summaries[site] = summary_df

    if not all_summaries:
        print("❌ 无可用数据，退出")
        return None

    # 生成合并汇总
    combined = pd.concat(all_summaries.values(), ignore_index=True)
    print(f"\n{'='*60}")
    print(f"  合并汇总 ({len(all_summaries)} 个站点)")
    print(f"{'='*60}")

    group_totals = combined.groupby("Site").agg(
        Items=("Item", "count"),
        Prev_Balance=("Prev_Balance", "sum"),
        Received_AMT=("Received_AMT", "sum"),
        Consumed_AMT=("Consumed_AMT", "sum"),
        Other_AMT=("Other_AMT", "sum"),
        Balance_AMT=("Balance_AMT", "sum"),
    ).reset_index()

    for _, row in group_totals.iterrows():
        s = row["Site"]
        print(f"\n  Site {s} ({SITE_NAMES.get(s, s)}):")
        print(f"    Items: {row['Items']:,}")
        print(f"    期初: ${row['Prev_Balance']:,.2f}  +Recv: ${row['Received_AMT']:,.2f}  "
              f"+Cons: ${row['Consumed_AMT']:,.2f}  +Other: ${row['Other_AMT']:,.2f}  "
              f"= ${row['Balance_AMT']:,.2f}")

    grand_prev = combined["Prev_Balance"].sum()
    grand_recv = combined["Received_AMT"].sum()
    grand_cons = combined["Consumed_AMT"].sum()
    grand_other = combined["Other_AMT"].sum()
    grand_bal = grand_prev + grand_recv + grand_cons + grand_other
    print(f"\n  🏢 ALL SITES Grand Total (USD):")
    print(f"    期初: ${grand_prev:,.2f}  +Recv: ${grand_recv:,.2f}  "
          f"+Cons: ${grand_cons:,.2f}  +Other: ${grand_other:,.2f}  "
          f"= ${grand_bal:,.2f}")

    return {
        "combined": combined,
        "group_totals": group_totals,
        "grand_prev": grand_prev, "grand_recv": grand_recv,
        "grand_cons": grand_cons, "grand_other": grand_other,
        "grand_bal": grand_bal, "out_dir": out_dir,
    }


# ──────────────────────────────────────────────────────────────
# 发送汇总邮件（SMTP 版本）
# ──────────────────────────────────────────────────────────────
def send_summary_email(
    result: dict,
    to_addr: str = "jason.pang@nai-group.com;shirley.ni@nai-group.com;devin.hua@nai-group.com;chn_planners@nai-group.com;chn_buyer@nai-group.com",
    cc_addr: str = "sky.li@nai-group.com;frank.liu@nai-group.com;shirley.ni@nai-group.com",
    smtp_host: str = "localhost",
    smtp_port: int = 25,
    smtp_user: str = "",
    smtp_password: str = "",
    smtp_tls: bool = False,
    from_addr: str = "inventory-report@nai-group.com",
):
    """生成固定格式的 HTML 邮件并通过 SMTP 发送"""
    gt = result["group_totals"]
    combined = result["combined"]

    def fmt_num(v):
        if v < 0:
            return f"-${abs(v):,.2f}"
        return f"${v:,.2f}"

    def proj_count(site):
        site_df = combined[combined["Site"] == site]
        pc = site_df["ProjectCode"].dropna()
        pc = pc[pc.str.strip() != ""]
        return len(pc.unique())

    # 构建表格行
    rows_html = ""
    for _, row in gt.iterrows():
        s = row["Site"]
        label = f"{s} ({SITE_NAMES.get(s, s)})"
        pc_count = proj_count(s)
        rows_html += (
            f"<tr>"
            f"<td style='text-align:left'>{label}</td>"
            f"<td>{int(row['Items']):,}</td>"
            f"<td>{pc_count}</td>"
            f"<td>{fmt_num(row['Prev_Balance'])}</td>"
            f"<td>{fmt_num(row['Received_AMT'])}</td>"
            f"<td>{fmt_num(row['Consumed_AMT'])}</td>"
            f"<td>{fmt_num(row['Other_AMT'])}</td>"
            f"<td><b>{fmt_num(row['Balance_AMT'])}</b></td>"
            f"</tr>\n"
        )

    total_items = int(gt["Items"].sum())
    total_pc = sum(proj_count(s) for s in gt["Site"])
    rows_html += (
        f"<tr style='background-color:#D6E4F0;font-weight:bold'>"
        f"<td style='text-align:left'>Grand Total</td>"
        f"<td>{total_items:,}</td><td>-</td>"
        f"<td>{fmt_num(result['grand_prev'])}</td>"
        f"<td>{fmt_num(result['grand_recv'])}</td>"
        f"<td>{fmt_num(result['grand_cons'])}</td>"
        f"<td>{fmt_num(result['grand_other'])}</td>"
        f"<td><b>{fmt_num(result['grand_bal'])}</b></td>"
        f"</tr>\n"
    )

    today = datetime.today()
    month_label = today.strftime("%B %Y")
    date_label = today.strftime("%B %d %Y")
    prev_month = (today.replace(day=1) - timedelta(days=1))
    prev_eom_label = prev_month.strftime("%m/%d")

    html_body = f"""<div style="font-family:Calibri,Arial,sans-serif;font-size:11pt">
<p>Below is the {month_label} Inventory Valuation Tracking Report (Purchased Materials, PMTCode=P). All amounts in USD. Site 330 CNY amounts converted at rate 6.838784.</p>

<table border="1" cellpadding="5" cellspacing="0"
  style="border-collapse:collapse;font-size:10pt;text-align:right">
<tr style="background-color:#4472C4;color:white;text-align:center">
  <th>Site</th><th>Items</th><th>Projects</th>
  <th>{prev_eom_label} Balance</th><th>MTD Received</th>
  <th>MTD Consumption</th><th>MTD Other Transaction</th>
  <th>MTD Daily Balance</th>
</tr>
{rows_html}
</table>
</div>"""

    # ── 构建 MIME 邮件 ──
    msg = MIMEMultipart()
    msg["From"] = from_addr
    msg["To"] = to_addr
    if cc_addr:
        msg["Cc"] = cc_addr
    msg["Subject"] = f"Inventory Valuation Tracking Daily Report - {date_label}"
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # 附件
    for site in gt["Site"]:
        fname = f"Inventory_Balance_{SITE_NAMES.get(site, site)}_{today.strftime('%Y%m')}.xlsx"
        fpath = str(result["out_dir"] / fname)
        if os.path.exists(fpath):
            try:
                with open(fpath, "rb") as f:
                    part = MIMEBase("application", "octet-stream")
                    part.set_payload(f.read())
                encoders.encode_base64(part)
                part.add_header("Content-Disposition", f'attachment; filename="{fname}"')
                msg.attach(part)
                print(f"  📎 Attached: {fname}")
            except Exception as e:
                print(f"  ⚠️ Failed to attach {fname}: {e}")
        else:
            print(f"  ⚠️ File not found: {fpath}")

    # 收件人列表（To + Cc 合并）
    all_recipients = [a.strip() for a in to_addr.split(";") if a.strip()]
    if cc_addr:
        all_recipients += [a.strip() for a in cc_addr.split(";") if a.strip()]
    all_recipients = list(set(all_recipients))  # 去重

    # ── 发送 ──
    try:
        if smtp_tls:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        else:
            server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)

        if smtp_user and smtp_password:
            server.login(smtp_user, smtp_password)

        server.sendmail(from_addr, all_recipients, msg.as_string())
        server.quit()
        print(f"  ✅ 邮件已通过 SMTP ({smtp_host}:{smtp_port}) 发送")
        print(f"     To : {to_addr}")
        if cc_addr:
            print(f"     Cc : {cc_addr}")
    except Exception as e:
        print(f"  ❌ 邮件发送失败：{e}")
        raise


# ──────────────────────────────────────────────────────────────
# 期初库存生成 — 从 Infor CSI API 获取库存快照
# ──────────────────────────────────────────────────────────────

# Infor CSI API 站点参数配置
# clmParam 格式：M,PMT,B,ABC,0,1,,,,,,,0,0,{site}
# 取 PMTCode=P（Purchased Material）的 ABC 全等级，只含有库存品
INFOR_API_BASE = "https://mingle-ionapi.inforcloudsuite.com"
INFOR_TENANT = "NAIGROUP_PRD"
INFOR_IDO = "SLItemCostingReport"
INFOR_REPORT_PROC = "Rpt_ItemCostingSp"
INFOR_PROPERTIES = "Item,Itemdesc,Units,Unitcost,Unitscost"

# 每个站点的 clmParam 及 MongooseConfig header
# 格式说明：M=制造 / PMT=采购类型 / B=选 / ABC=ABC分类 / 0,1=参数占位 / 最后参数=SiteRef
SITE_API_CONFIG = {
    "310": {
        "clmParam": "M,PMT,B,ABC,0,1,,,,,,,0,0,310",
        "mongoose_config": "NAIGROUP_PRD_310",
    },
    "330": {
        "clmParam": "M,PMT,B,ABC,0,1,,,,,,,0,0,330",
        "mongoose_config": "NAIGROUP_PRD_330",
    },
    "410": {
        "clmParam": "M,PMT,B,ABC,0,1,,,,,,,0,0,410",
        "mongoose_config": "NAIGROUP_PRD_410",
    },
}


# ── Infor OAuth2 Token 缓存（模块级，进程生命周期内有效）────────
_infor_token_cache: str | None = None
_infor_token_expires_at: float = 0.0


def _read_infor_config() -> dict:
    """
    从环境变量读取 OAuth2 配置。
    优先级：环境变量 > 代码内默认值
    """
    def _env(env_key, default=""):
        return os.environ.get(env_key, "").strip() or default

    return {
        "token_url":    _env("INFOR_TOKEN_URL",
                             f"https://mingle-sso.inforcloudsuite.com:443/{INFOR_TENANT}/as/token.oauth2"),
        "auth_basic":   _env("INFOR_AUTH_BASIC"),
        "username":     _env("INFOR_USERNAME"),
        "password":     _env("INFOR_PASSWORD"),
        "bearer_token": _env("INFOR_BEARER_TOKEN"),
    }


def _fetch_oauth_token(config: dict) -> tuple[str, int]:
    """
    发起 OAuth2 token 请求，返回 (access_token, expires_in)。
    使用 urllib（不依赖 httpx），兼容 Docker 内的精简环境。
    """
    token_url = config["token_url"]
    auth_basic = config["auth_basic"]
    username = config["username"]
    password = config["password"]

    body = urllib.parse.urlencode({
        "grant_type": "password",
        "username": username,
        "password": password,
    }).encode("utf-8")

    req = urllib.request.Request(
        token_url,
        data=body,
        headers={
            "Authorization": f"Basic {auth_basic}",
            "Content-Type": "application/x-www-form-urlencoded",
        },
        method="POST",
    )

    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            raw = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(
            f"❌ OAuth2 Token 请求失败 (HTTP {e.code})\n"
            f"   URL: {token_url}\n"
            f"   响应: {err_body[:500]}"
        ) from e
    except urllib.error.URLError as e:
        raise RuntimeError(
            f"❌ Token 端点网络连接失败: {e.reason}\n"
            f"   URL: {token_url}\n"
            f"   请检查 VPN 或网络连接"
        ) from e

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(
            f"❌ Token 端点返回非 JSON: {e}\n"
            f"   响应: {raw[:500]}"
        ) from e

    token = data.get("access_token")
    if not token:
        error = data.get("error", "unknown")
        desc = data.get("error_description", "")
        raise RuntimeError(
            f"❌ Token 响应中无 access_token\n"
            f"   error: {error}\n"
            f"   description: {desc}"
        )

    expires_in = int(data.get("expires_in", 3600))
    return token, expires_in


def _load_infor_token(force_refresh: bool = False) -> str:
    """
    获取 Infor CSI API 的有效 access_token。

    策略：
      1. 如果有缓存 token 且未过期（提前 60s），直接返回
      2. 读取 OAuth2 配置（Basic Auth + password grant）
      3. POST token endpoint 获取新 token
      4. 对 400/429/502/503/504 瞬时错误自动重试 3 次
      5. 如果 OAuth2 配置缺失，降级使用手动 Bearer Token

    force_refresh=True 时忽略缓存，强制重新获取（用于 401 重试场景）。
    """
    global _infor_token_cache, _infor_token_expires_at

    # ── 检查缓存 ──
    if not force_refresh and _infor_token_cache:
        now = time.time()
        if now < _infor_token_expires_at - 60:
            return _infor_token_cache

    # ── 读取配置 ──
    config = _read_infor_config()

    # ── 模式 A：OAuth2 password grant（推荐，自动刷新）──
    if config["auth_basic"] and config["username"] and config["password"]:
        print(f"  🔐 OAuth2 获取 Token: {config['token_url']}")
        last_err = None
        for attempt in range(1, 4):
            try:
                token, expires_in = _fetch_oauth_token(config)
                _infor_token_cache = token
                _infor_token_expires_at = time.time() + expires_in
                print(f"  ✅ Token 获取成功，有效期 {expires_in}s")
                return token
            except urllib.error.HTTPError as e:
                last_err = e
                if e.code in (400, 429, 502, 503, 504):
                    wait = 2 ** (attempt - 1)
                    print(f"  ⚠️  Token 请求第 {attempt} 次失败 (HTTP {e.code})，{wait}s 后重试...")
                    time.sleep(wait)
                else:
                    raise
            except RuntimeError as e:
                last_err = e
                if attempt < 3:
                    wait = 2 ** (attempt - 1)
                    print(f"  ⚠️  Token 请求第 {attempt} 次异常：{e}")
                    print(f"  ⏳ {wait}s 后重试...")
                    time.sleep(wait)
                else:
                    break

        raise RuntimeError(f"❌ OAuth2 Token 连续获取失败（3次重试）：{last_err}") from last_err

    # ── 模式 B：手动 Bearer Token（降级，需要手动刷新）──
    if config["bearer_token"]:
        _infor_token_cache = config["bearer_token"]
        _infor_token_expires_at = time.time() + 86400  # 手动 token 假设 24h 有效
        print("  🔑 使用手动 Bearer Token（注意：过期需手动更新）")
        return _infor_token_cache

    raise RuntimeError(
        "❌ 未找到 Infor API 认证凭据！\n"
        "   请在 .env 文件中配置 OAuth2 凭据：\n"
        "   ── 推荐：OAuth2 password grant（自动刷新，无需手动维护 Token）──\n"
        "   INFOR_AUTH_BASIC=<Base64(client_id:client_secret)>\n"
        "   INFOR_USERNAME=<infor_username>\n"
        "   INFOR_PASSWORD=<infor_password>\n"
        "   ── 备选：手动 Bearer Token（需手动更新，不推荐）──\n"
        "   INFOR_BEARER_TOKEN=<token>\n"
        "   ── 获取 OAuth2 凭据 ──\n"
        "   1. client_id / client_secret：联系 Infor 管理员或从 csi_datawarehouse .env 获取\n"
        "   2. INFOR_AUTH_BASIC = Base64(client_id:client_secret)\n"
        "      Linux: echo -n 'client_id:client_secret' | base64\n"
        "      Python: base64.b64encode(b'client_id:client_secret').decode()\n"
    )


def _fetch_infor_site(site_ref: str, token: str, token_expired_retry: bool = False) -> pd.DataFrame:
    """
    调用 Infor CSI IDO API 获取指定站点的库存成本报告（Purchased Material）。
    返回 DataFrame，列：Item, Itemdesc, Prodcode, Units, Unitcost

    token_expired_retry=True 时表示已在重试中，不再重试 401。
    """
    site_cfg = SITE_API_CONFIG[site_ref]
    clm_param = urllib.parse.quote(site_cfg["clmParam"], safe="")
    url = (
        f"{INFOR_API_BASE}/{INFOR_TENANT}/CSI/IDORequestService/ido/load/{INFOR_IDO}"
        f"?clm={INFOR_REPORT_PROC}"
        f"&properties={INFOR_PROPERTIES}"
        f"&clmParam={clm_param}"
    )

    headers = {
        "Authorization": f"Bearer {token}",
        "X-Infor-MongooseConfig": site_cfg["mongoose_config"],
        "Accept": "application/json",
        "Content-Type": "application/json",
    }

    print(f"  🌐 调用 Infor API: Site {site_ref} ({site_cfg['mongoose_config']})...")

    req = urllib.request.Request(url, headers=headers, method="GET")
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            raw = resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="replace")
        if e.code == 401 and not token_expired_retry:
            # Token 过期 → 强制刷新后重试一次
            print(f"  🔄 Site {site_ref}: Token 过期 (401)，强制刷新...")
            new_token = _load_infor_token(force_refresh=True)
            return _fetch_infor_site(site_ref, new_token, token_expired_retry=True)
        elif e.code == 401:
            raise RuntimeError(
                f"❌ Infor API 认证失败 (401) - Site {site_ref}\n"
                f"   Token 刷新后仍然无效，请检查 OAuth2 凭据"
            ) from e
        elif e.code == 502:
            raise RuntimeError(
                f"❌ Infor API 502 Bad Gateway - Site {site_ref}\n"
                f"   可能原因：VPN 未连接、Infor CloudSuite 服务暂时不可用\n"
                f"   响应：{body[:300]}"
            ) from e
        else:
            raise RuntimeError(
                f"❌ Infor API HTTP {e.code} - Site {site_ref}\n"
                f"   响应：{body[:300]}"
            ) from e
    except urllib.error.URLError as e:
        raise RuntimeError(
            f"❌ 网络连接失败 - Site {site_ref}: {e.reason}\n"
            f"   请检查 VPN 或网络连接"
        ) from e

    # 解析 JSON
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(
            f"❌ Infor API 返回非 JSON 内容 - Site {site_ref}: {e}\n"
            f"   原始响应前 500 字符：{raw[:500]}"
        ) from e

    # IDO 响应格式：{"Items": {"Items": [{"PropValue": [v1, v2, ...]}, ...]}}
    # 或：{"Items": [{"PropValue": [v1, v2, ...]}, ...]}
    rows = []
    props = INFOR_PROPERTIES.split(",")

    try:
        item_list = data.get("Items", {})
        if isinstance(item_list, dict):
            item_list = item_list.get("Items", [])
        if not isinstance(item_list, list):
            raise ValueError(f"响应结构异常，Items 不是列表: {type(item_list)}")

        for record in item_list:
            values = record.get("PropValue", [])
            if len(values) < len(props):
                values += [""] * (len(props) - len(values))
            row = dict(zip(props, values))
            rows.append(row)

    except Exception as e:
        raise RuntimeError(
            f"❌ Infor API 响应解析失败 - Site {site_ref}: {e}\n"
            f"   响应结构：{str(data)[:500]}"
        ) from e

    df = pd.DataFrame(rows, columns=props)
    print(f"  ✅ Site {site_ref}: API 返回 {len(df)} 条记录")

    if df.empty:
        print(f"  ⚠️  Site {site_ref}: 未返回数据（可能 clmParam 参数有误或站点无采购物料）")
        return pd.DataFrame(columns=["Item", "Description", "Per", "Unitcost", "Unitscost"])

    # 列名标准化
    # Units   → Per       (库存数量)
    # Unitcost → Unitcost  (标准单价，保留原名)
    # Unitscost → Unitscost (扩展金额 = Units × Unitcost，API 直接提供，无需计算)
    df = df.rename(columns={
        "Itemdesc":  "Description",
        "Units":     "Per",
    })

    df["Item"]      = df["Item"].astype(str).str.strip().str.upper()
    df["Description"] = df["Description"].astype(str).str.strip()
    df["Per"]       = pd.to_numeric(df["Per"],       errors="coerce").round(8).fillna(0)
    df["Unitcost"]  = pd.to_numeric(df["Unitcost"],  errors="coerce").round(8).fillna(0)
    df["Unitscost"] = pd.to_numeric(df["Unitscost"], errors="coerce").round(8).fillna(0)

    # 只保留有扩展金额的行（Unitscost != 0）
    df = df[df["Unitscost"] != 0].copy()

    return df[["Item", "Description", "Per", "Unitcost", "Unitscost"]]


def _db_connect(server, username, password, database, port=1433):
    """独立的数据库连接函数（供 generate_opening_balance SQL fallback 使用）"""
    host = server
    instance = None
    if "\\" in host:
        host, instance = host.split("\\", 1)

    kwargs = dict(
        server=f"{host}\\{instance}" if instance else host,
        user=username,
        password=password,
        database=database,
        tds_version="7.4",
        login_timeout=15,
        conn_properties="SET TEXTSIZE 65536",
    )
    if not instance:
        kwargs["port"] = port

    return pymssql.connect(**kwargs)


def _enrich_with_slitems(df: pd.DataFrame, site_ref: str,
                          server, username, password, database, port=1433) -> pd.DataFrame:
    """
    从 SLItems 表按 Item + SiteRef 关联补充项目信息。
    补充字段：ProductCode, Sourcing（PMTCode）
    若数据库不可用，则以空值填充（不阻断主流程）。
    """
    if df.empty:
        return df

    items = df["Item"].dropna().unique().tolist()
    if not items:
        return df

    try:
        conn = _db_connect(server, username, password, database, port)
        placeholders = ",".join(["%s"] * len(items))
        query = f"""
            SELECT UPPER(LTRIM(RTRIM(item))) AS Item,
                   ProductCode,
                   PMTCode AS Sourcing
            FROM [csi_datawarehouse].[dbo].[SLItems]
            WHERE SiteRef = %s
              AND UPPER(LTRIM(RTRIM(item))) IN ({placeholders})
        """
        params = [site_ref] + items
        ref_df = pd.read_sql(query, conn, params=params)
        conn.close()

        ref_df["Item"] = ref_df["Item"].astype(str).str.strip().str.upper()
        ref_df["ProductCode"] = ref_df["ProductCode"].astype(str).str.strip()
        ref_df["Sourcing"] = ref_df["Sourcing"].astype(str).str.strip()
        ref_df = ref_df.drop_duplicates(subset="Item")

        df = df.merge(ref_df[["Item", "ProductCode", "Sourcing"]], on="Item", how="left")
        df["ProductCode"] = df["ProductCode"].fillna("")
        df["Sourcing"] = df["Sourcing"].fillna("Unknown")
        print(f"  🔗 Site {site_ref}: SLItems 关联成功，{len(ref_df)} 条记录匹配")
    except Exception as e:
        print(f"  ⚠️  Site {site_ref}: SLItems 关联失败（{e}），ProductCode/Sourcing 将为空")
        df["ProductCode"] = ""
        df["Sourcing"] = ""

    return df


def _fetch_opening_via_sql(server, username, password, database, port=1433) -> pd.DataFrame:
    """
    SQL Fallback：直接查询 SLItems 表获取库存快照。
    仅在 Infor API 不可用时调用。
    """
    conn = _db_connect(server, username, password, database, port)
    query = """
        SELECT SiteRef,
               item,
               Description,
               'Purchased' AS Sourcing,
               ProductCode,
               CAST(OnHand AS DECIMAL(20,8)) AS Per,
               CAST(DerUnitCost AS DECIMAL(20,8)) AS Unitcost,
               CAST(OnHand * DerUnitCost AS DECIMAL(20,8)) AS Unitscost
        FROM [csi_datawarehouse].[dbo].[SLItems]
        WHERE SiteRef IN ('310', '330', '410')
          AND PMTCode = 'P'
          AND OnHand <> 0
        ORDER BY SiteRef, ProductCode, item
    """
    df = pd.read_sql(query, conn)
    conn.close()
    return df


def generate_opening_balance(
    server=None, username=None, password=None, database=None, port=1433,
    output_dir="Previous Balance",
    use_api: bool = True,
):
    """
    生成期初库存余额 Excel 文件（一个站点一个文件）。
    文件写入 output_dir/site XXX.xlsx，供日常报表读取。
    仅含 PMTCode='P'（采购物料），与日常报表口径一致。

    数据来源优先级：
      1. Infor CSI API（use_api=True，默认）：调用 SLItemCostingReport IDO
      2. SQL Fallback（use_api=False 或 API 失败时）：直接查询 SLItems 表

    参数：
      server/username/password/database/port — SQL fallback 用的数据库连接参数
      output_dir — Excel 文件输出目录
      use_api    — 是否优先使用 Infor CSI API（默认 True）
    """
    print("=" * 60)
    print("  生成期初库存余额文件")
    print("=" * 60)

    out_path = Path(output_dir)
    out_path.mkdir(parents=True, exist_ok=True)

    site_info = {"310": "Plant1", "330": "Plant2", "410": "PNG"}
    sites = ["310", "330", "410"]

    # ── 尝试从 Infor CSI API 获取数据 ──
    if use_api:
        try:
            token = _load_infor_token()
        except RuntimeError as e:
            print(f"  {e}")
            print("  ⚠️  将降级使用 SQL 直连方式...")
            use_api = False

    if use_api:
        print("  📡 数据来源：Infor CSI API")
        results = []
        api_failed_sites = []

        for site_ref in sites:
            try:
                df = _fetch_infor_site(site_ref, token)
            except RuntimeError as e:
                print(f"  ❌ Site {site_ref} API 调用失败：{e}")
                api_failed_sites.append(site_ref)
                continue

            if df.empty:
                api_failed_sites.append(site_ref)
                continue

            # 从 SLItems 关联 ProductCode / Sourcing
            if server:
                df = _enrich_with_slitems(df, site_ref, server, username, password, database, port)
            else:
                df["ProductCode"] = ""
                df["Sourcing"] = ""

            _write_opening_excel(df, site_ref, out_path, site_info, results)

        # 对 API 失败的站点尝试 SQL fallback
        if api_failed_sites and server:
            print(f"\n  ↩️  以下站点将使用 SQL fallback：{api_failed_sites}")
            try:
                fallback_df = _fetch_opening_via_sql(server, username, password, database, port)
                for site_ref in api_failed_sites:
                    site_df = fallback_df[fallback_df["SiteRef"] == site_ref].copy()
                    site_df = site_df[["item", "Description", "Per", "Unitcost", "Unitscost", "ProductCode", "Sourcing"]].copy()
                    site_df.columns = ["Item", "Description", "Per", "Unitcost", "Unitscost", "ProductCode", "Sourcing"]
                    site_df["Item"] = site_df["Item"].astype(str).str.strip().str.upper()
                    _write_opening_excel(site_df, site_ref, out_path, site_info, results)
            except Exception as e:
                print(f"  ❌ SQL fallback 也失败：{e}")
        elif api_failed_sites:
            print(f"  ⚠️  以下站点无法获取数据（无 SQL fallback 配置）：{api_failed_sites}")

    else:
        # 纯 SQL 模式
        print("  🗄️  数据来源：SQL Server SLItems 表")
        if not server:
            print("  ❌ SQL 模式需要数据库连接参数（server/username/password/database）")
            return None

        try:
            fallback_df = _fetch_opening_via_sql(server, username, password, database, port)
        except Exception as e:
            print(f"  ❌ SQL 查询失败：{e}")
            return None

        results = []
        for site_ref in sites:
            site_df = fallback_df[fallback_df["SiteRef"] == site_ref].copy()
            site_df = site_df[["item", "Description", "Per", "Unitcost", "Unitscost", "ProductCode", "Sourcing"]].copy()
            site_df.columns = ["Item", "Description", "Per", "Unitcost", "Unitscost", "ProductCode", "Sourcing"]
            site_df["Item"] = site_df["Item"].astype(str).str.strip().str.upper()
            _write_opening_excel(site_df, site_ref, out_path, site_info, results)

    if not results:
        print("  ⚠️ 所有站点均未能获取数据")
        return None

    grand = sum(r["value"] for r in results)
    print(f"\n  📊 总计: {sum(r['items'] for r in results)} items, Grand Total: ${grand:,.2f}")
    print(f"  📁 文件目录: {out_path.resolve()}")
    return results


def _write_opening_excel(df: pd.DataFrame, site_ref: str, out_path: Path,
                          site_info: dict, results: list) -> None:
    """将单站点期初数据写入 Excel 并追加到 results 列表"""
    label = site_info.get(site_ref, site_ref)
    fname = out_path / f"site {site_ref}.xlsx"

    # 数值列统一保留 8 位小数
    for col in ["Per", "Unitcost", "Unitscost"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce").round(8)

    df.to_excel(fname, index=False, sheet_name="Opening Balance")

    total_items = len(df)
    # 直接用 Unitscost（扩展金额）列求和；fallback 到 Per × Unitcost
    if "Unitscost" in df.columns:
        total_value = pd.to_numeric(df["Unitscost"], errors="coerce").fillna(0).sum()
    else:
        total_value = (
            pd.to_numeric(df.get("Per", 0), errors="coerce").fillna(0) *
            pd.to_numeric(df.get("Unitcost", 0), errors="coerce").fillna(0)
        ).sum()

    print(f"  ✅ Site {site_ref} ({label}): {total_items} items, 金额: ${total_value:,.2f}")
    print(f"     → {fname}")
    results.append({"site": site_ref, "label": label, "items": total_items, "value": total_value})


# ──────────────────────────────────────────────────────────────
# 守护进程 — 每月1日 00:15 生成期初 + 每天 09:00 跑报表
# ──────────────────────────────────────────────────────────────
DAEMON_HOUR_OPENING = 0
DAEMON_MINUTE_OPENING = 15
DAEMON_HOUR_REPORT = 9
DAEMON_MINUTE_REPORT = 0

def _next_schedule(now):
    """
    计算下一次执行时间和任务类型。
    规则：
      - 每月1日 00:15：生成期初库存文件
      - 每天 09:00：运行库存跟踪报表
    返回 (target_datetime, task_type)  task_type = 'opening' | 'report'
    """
    def _make(hour, minute):
        return now.replace(hour=hour, minute=minute, second=0, microsecond=0)

    today_opening = _make(DAEMON_HOUR_OPENING, DAEMON_MINUTE_OPENING)
    today_report = _make(DAEMON_HOUR_REPORT, DAEMON_MINUTE_REPORT)

    candidates = []
    # 每月1日 00:15 生成期初
    if now.day == 1 and now < today_opening:
        candidates.append((today_opening, "opening"))
    # 每天 09:00 跑报表
    if now < today_report:
        candidates.append((today_report, "report"))

    if candidates:
        return min(candidates, key=lambda x: x[0])

    # 今天的任务都过了，看明天
    tomorrow = (now + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    if tomorrow.day == 1:
        return (tomorrow.replace(hour=DAEMON_HOUR_OPENING, minute=DAEMON_MINUTE_OPENING), "opening")
    else:
        return (tomorrow.replace(hour=DAEMON_HOUR_REPORT, minute=DAEMON_MINUTE_REPORT), "report")


def schedule_loop(run_func, args):
    """
    双调度守护进程：
      - 每月1日 00:15：自动生成期初库存文件（优先 Infor CSI API，降级 SQL）
      - 每天 09:00：运行库存跟踪报表 + 发邮件
    Docker 停止时收到 SIGTERM 自然退出。
    """
    gen_args = dict(
        server=args.server, username=args.username,
        password=args.password, database=args.database,
        port=args.db_port, output_dir=args.prev_dir,
        use_api=True,
    )

    print(f"⏰ Daemon mode started:")
    print(f"   每月1日 {DAEMON_HOUR_OPENING:02d}:{DAEMON_MINUTE_OPENING:02d} → 生成期初库存文件")
    print(f"   每天   {DAEMON_HOUR_REPORT:02d}:{DAEMON_MINUTE_REPORT:02d} → 运行库存跟踪报表")

    while True:
        target, task_type = _next_schedule(datetime.now())
        wait_seconds = (target - datetime.now()).total_seconds()

        label = "生成期初库存" if task_type == "opening" else "运行库存报表"
        print(f"⏳ Next: {target.strftime('%Y-%m-%d %H:%M:%S')} [{label}] ({wait_seconds/3600:.1f}h)")

        try:
            time.sleep(max(wait_seconds, 0))
        except KeyboardInterrupt:
            print("\n🛑 Daemon stopped.")
            break

        try:
            if task_type == "opening":
                print(f"\n{'='*60}")
                print(f"  🔄 定时任务：生成期初库存文件")
                print(f"{'='*60}")
                generate_opening_balance(**gen_args)
                # 生成期初后，当天 09:00 还会自动跑报表
            else:
                run_func(args)
        except Exception as e:
            print(f"❌ Scheduled run failed: {e}")
            import traceback
            traceback.print_exc()
            time.sleep(300)


def run_once(args):
    """执行一次完整的库存跟踪+邮件流程"""
    if args.all_sites:
        result = run_all_sites(
            server=args.server, database=args.database,
            username=args.username, password=args.password, port=args.db_port,
            prev_dir=args.prev_dir, output_dir=args.output_dir,
        )
        if result and not args.no_email:
            print("\n📧 发送汇总邮件...")
            send_summary_email(
                result,
                to_addr=args.email_to,
                cc_addr=args.email_cc,
                smtp_host=args.smtp_host,
                smtp_port=args.smtp_port,
                smtp_user=args.smtp_user,
                smtp_password=args.smtp_password,
                smtp_tls=args.smtp_tls,
                from_addr=args.email_from,
            )
        return result
    else:
        if not args.prev_file:
            prev_dir = Path(args.prev_dir)
            site_label = SITE_NAMES.get(args.site, args.site)
            for f in prev_dir.iterdir():
                if site_label.upper() in f.name.upper() or args.site in f.name:
                    args.prev_file = str(f)
                    break
            if not args.prev_file:
                print(f"❌ 未找到 Site {args.site} 的期初余额文件。请用 -f 指定。")
                sys.exit(1)

        tracker = InventoryTracker(
            server=args.server, database=args.database,
            username=args.username, password=args.password, port=args.db_port,
            site_ref=args.site, prev_balance_file=args.prev_file,
            output_file=args.output,
        )
        return tracker.run()


# ──────────────────────────────────────────────────────────────
# CLI
# ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="库存金额跟踪系统（pymssql + SMTP）",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用示例:
  # 单次批量运行
  python inventory_tracking.py --all-sites

  # 守护模式：每天 09:00 自动执行，启动后不立即跑（Docker 推荐）
  python inventory_tracking.py --all-sites --daemon

  # 单站点运行
  python inventory_tracking.py --site 310 -f "Previous Balance/site 310.xlsx"

  # 不发送邮件
  python inventory_tracking.py --all-sites --no-email
        """,
    )
    # 数据库
    parser.add_argument("-s", "--server",   default=os.environ.get("SQL_SERVER_HOST",     r"SUZVPRINT01\CUSTOMSSYS"))
    parser.add_argument("-d", "--database", default=os.environ.get("SQL_SERVER_DATABASE", "csi_datawarehouse"))
    parser.add_argument("-u", "--username", default=os.environ.get("SQL_SERVER_USERNAME", "sa"))
    parser.add_argument("-p", "--password", default=os.environ.get("SQL_SERVER_PASSWORD", ""))
    parser.add_argument("--db-port", type=int, default=int(os.environ.get("SQL_SERVER_PORT", 1433)))
    # 模式
    parser.add_argument("--site", default="310", help="单站点模式 (默认310)")
    parser.add_argument("-f", "--prev-file", help="单站点期初余额Excel路径")
    parser.add_argument("--all-sites", action="store_true", help="批量运行310/330/410")
    parser.add_argument("--daemon", action="store_true", help="守护模式：每月1日 00:15 生成期初 + 每天 09:00 跑报表")
    parser.add_argument("--generate-opening", action="store_true", help="手动生成期初库存文件（优先 Infor CSI API，降级 SQL）")
    parser.add_argument("--no-api", action="store_true", help="生成期初时强制使用 SQL 模式（跳过 Infor CSI API）")
    parser.add_argument("--prev-dir", default=os.environ.get("PREV_DIR", "Previous Balance"))
    parser.add_argument("--output-dir", default=os.environ.get("OUTPUT_DIR", None))
    parser.add_argument("-o", "--output", help="输出文件名 (单站点模式)")
    # 邮件
    parser.add_argument("--no-email", action="store_true", help="不发送邮件")
    parser.add_argument("--email-to",
        default=os.environ.get("MAIL_TO", "jason.pang@nai-group.com;shirley.ni@nai-group.com;devin.hua@nai-group.com;chn_planners@nai-group.com;chn_buyer@nai-group.com"))
    parser.add_argument("--email-cc",
        default=os.environ.get("MAIL_CC", "sky.li@nai-group.com;frank.liu@nai-group.com;shirley.ni@nai-group.com"))
    parser.add_argument("--email-from",
        default=os.environ.get("MAIL_FROM", "suzinventoryvaluationdailyreport@nai-group.com"))
    # SMTP（优先级：.env → 默认值）
    parser.add_argument("--smtp-host",     default=os.environ.get("SMTP_HOST", "localhost"))
    parser.add_argument("--smtp-port",     type=int, default=int(os.environ.get("SMTP_PORT", 25)))
    parser.add_argument("--smtp-user",     default=os.environ.get("SMTP_USER", ""))
    parser.add_argument("--smtp-password", default=os.environ.get("SMTP_PASSWORD", ""))
    parser.add_argument("--smtp-tls",      action="store_true",
        default=os.environ.get("SMTP_TLS", "false").lower() == "true")

    args = parser.parse_args()

    if args.generate_opening:
        # 手动生成期初模式（优先 Infor CSI API，降级 SQL）
        generate_opening_balance(
            server=args.server, username=args.username,
            password=args.password, database=args.database,
            port=args.db_port, output_dir=args.prev_dir,
            use_api=not args.no_api,
        )
    elif args.daemon:
        # 守护模式：双调度（每月1日 00:15 期初 + 每天 09:00 报表）
        schedule_loop(lambda a: run_once(a), args)
    else:
        run_once(args)


if __name__ == "__main__":
    # ── 加载 .env 到环境变量 ──
    try:
        from dotenv import load_dotenv
        load_dotenv()
    except ImportError:
        pass
    main()
