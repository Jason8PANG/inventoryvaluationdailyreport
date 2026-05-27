from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "PrevBalance"

# Title
ws.merge_cells("A1:C1")
ws["A1"] = "期初库存余额模板 (Previous Balance Template)"
ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
ws["A1"].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Headers
headers = ["Item", "Prev_Qty", "Prev_Balance"]
for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col_num)
    cell.value = header
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center")
    cell.border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

# Sample data
sample_data = [
    ["ITM-001", 100, 5000.00],
    ["ITM-002", 50, 2500.00],
    ["ITM-003", 200, 12000.00],
]

for row_idx, row_data in enumerate(sample_data, start=4):
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.value = value
        cell.border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        if col_num > 1:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0.00"

# Column widths
ws.column_dimensions["A"].width = 20
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 18

# Add instruction note
ws.merge_cells("A6:C6")
ws["A6"] = "说明：请替换为实际上月底库存数据，保留 Item / Prev_Qty / Prev_Balance 三列"
ws["A6"].font = Font(italic=True, color="666666")

output_path = "prev_balance_sample.xlsx"
wb.save(output_path)
print(f"Sample file created: {output_path}")
