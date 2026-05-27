#!/usr/bin/env python3
"""
库存金额跟踪系统 - 演示/测试版本
==============================
此版本不连接真实数据库，使用模拟数据演示分类和计算逻辑。
用于验证分类规则是否正确，无需安装 pyodbc。
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── TransType + RefType → Category 映射 ──
CATEGORY_MAP = {
    ("R", "P"): "Received",
    ("W", "P"): "Received",
    ("I", "J"): "Consumed",
    ("W", "J"): "Consumed",
    ("S", "O"): "Consumed",
    ("A", "I"): "Other",
    ("M", "I"): "Other",
    ("G", "I"): "Other",
    ("H", "I"): "Other",
    ("C", "J"): "Other",
    ("F", "J"): "Other",
    ("N", "J"): "Other",
    ("W", "R"): "Other",
}

# ── 模拟数据 ──
print("🧪 生成模拟测试数据...")

# 模拟期初余额
prev_balance_data = {
    "Item": ["ITM-001", "ITM-002", "ITM-003"],
    "Prev_Qty": [100, 50, 200],
    "Prev_Balance": [5000.00, 2500.00, 12000.00],
}
prev_df = pd.DataFrame(prev_balance_data)

# 模拟 MTD 事务
trans_data = [
    # Item, TransDate, TransType, RefType, Qty, TotalAmt
    ["ITM-001", "2026-05-05", "R", "P", 50, 2500.00],   # PO Receipt → Received
    ["ITM-001", "2026-05-10", "I", "J", -20, -1000.00], # Job Issue → Consumed
    ["ITM-001", "2026-05-15", "A", "I", 5, 250.00],    # Adjustment → Other
    ["ITM-002", "2026-05-03", "R", "P", 30, 1500.00],   # PO Receipt → Received
    ["ITM-002", "2026-05-08", "S", "O", -15, -750.00],  # Order Ship → Consumed
    ["ITM-002", "2026-05-12", "M", "I", -10, -500.00],  # Stock Move → Other
    ["ITM-003", "2026-05-02", "W", "P", 25, 1500.00],   # PO Withdraw → Received
    ["ITM-003", "2026-05-07", "W", "J", -30, -1800.00], # Job Withdrawal → Consumed
    ["ITM-003", "2026-05-14", "C", "J", -5, -300.00],   # Job Complete → Other
]

trans_df = pd.DataFrame(trans_data, columns=[
    "Item", "TransDate", "TransType", "RefType", "Qty", "TotalAmt"
])

print(f"✅ 期初余额：{len(prev_df)} 个 Item")
print(f"✅ MTD 事务：{len(trans_df)} 条记录")

# ── 分类 ──
trans_df["Category"] = trans_df.apply(
    lambda r: CATEGORY_MAP.get((r["TransType"], r["RefType"]), "Other"), axis=1
)

print("\n📋 事务分类明细：")
for _, row in trans_df.iterrows():
    print(f"   {row['Item']} | {row['TransDate']} | {row['TransType']}/{row['RefType']} | "
          f"{row['Category']:>10} | Qty:{row['Qty']:>6.2f} | Amt:{row['TotalAmt']:>8.2f}")

# ── 按 Item 汇总 ──
summary = {}
for item in prev_df["Item"]:
    prev_row = prev_df[prev_df["Item"] == item].iloc[0]
    summary[item] = {
        "Prev_Qty": prev_row["Prev_Qty"],
        "Prev_Balance": prev_row["Prev_Balance"],
        "Received_Qty": 0, "Received_AMT": 0,
        "Consumed_Qty": 0, "Consumed_AMT": 0,
        "Other_Qty": 0, "Other_AMT": 0,
    }

for _, row in trans_df.iterrows():
    item = row["Item"]
    cat = row["Category"]
    qty = row["Qty"]
    amt = row["TotalAmt"]

    if item not in summary:
        summary[item] = {
            "Prev_Qty": 0, "Prev_Balance": 0,
            "Received_Qty": 0, "Received_AMT": 0,
            "Consumed_Qty": 0, "Consumed_AMT": 0,
            "Other_Qty": 0, "Other_AMT": 0,
        }

    if cat == "Received":
        summary[item]["Received_Qty"] += qty
        summary[item]["Received_AMT"] += amt
    elif cat == "Consumed":
        summary[item]["Consumed_Qty"] += qty
        summary[item]["Consumed_AMT"] += amt
    else:
        summary[item]["Other_Qty"] += qty
        summary[item]["Other_AMT"] += amt

# ── 计算期末余额 ──
print("\n📊 汇总计算结果：")
print("=" * 100)
print(f"{'Item':<12} {'期初金额':>12} {'Received':>12} {'Consumed':>12} {'Other':>12} {'期末余额':>12}")
print("-" * 100)

for item, data in sorted(summary.items()):
    balance = data["Prev_Balance"] + data["Received_AMT"] - data["Consumed_AMT"] - data["Other_AMT"]
    net_qty = data["Prev_Qty"] + data["Received_Qty"] - data["Consumed_Qty"] - data["Other_Qty"]
    data["Balance_AMT"] = balance
    data["Balance_Qty"] = net_qty
    print(f"{item:<12} {data['Prev_Balance']:>12.2f} {data['Received_AMT']:>12.2f} "
          f"{data['Consumed_AMT']:>12.2f} {data['Other_AMT']:>12.2f} {balance:>12.2f}")

print("=" * 100)
print("\n💡 验证公式：期末余额 = 期初 + Received - Consumed - Other")

# ── 导出 Excel ──
print("\n📤 导出演示 Excel...")
wb = Workbook()

# Summary sheet
ws = wb.active
ws.title = "Summary"

ws.merge_cells("A1:K1")
ws["A1"] = "库存金额跟踪报表 - 演示数据"
ws["A1"].font = Font(bold=True, size=14)
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:K2")
ws["A2"] = f"报告周期：{datetime.today().strftime('%Y年%m月')} (演示用模拟数据)"
ws["A2"].alignment = Alignment(horizontal="center")
ws.append([])

headers = ["Item", "期初数量", "期初金额", "Received Qty", "Received AMT",
           "Consumed Qty", "Consumed AMT", "Other Qty", "Other AMT",
           "期末数量", "期末金额 (Balance)"]
ws.append(headers)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

for col_num, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

row_idx = 5
for item, data in sorted(summary.items()):
    ws.append([
        item,
        data["Prev_Qty"], data["Prev_Balance"],
        data["Received_Qty"], data["Received_AMT"],
        data["Consumed_Qty"], data["Consumed_AMT"],
        data["Other_Qty"], data["Other_AMT"],
        data["Balance_Qty"], data["Balance_AMT"],
    ])
    for col_num in range(1, 12):
        cell = ws.cell(row=row_idx, column=col_num)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="right" if col_num > 1 else "left")
        if col_num == 11:
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            cell.font = Font(bold=True)
    row_idx += 1

col_widths = [15, 12, 15, 14, 14, 14, 14, 12, 12, 12, 18]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[chr(64 + i)].width = w

# Detail sheet
ws_det = wb.create_sheet("Detail")
ws_det.append(["物料事务明细"])
ws_det.merge_cells("A1:F1")
ws_det["A1"].font = Font(bold=True, size=14)
ws_det["A1"].alignment = Alignment(horizontal="center")
ws_det.append([])

detail_headers = ["Item", "TransDate", "TransType", "RefType", "Category", "Qty", "TotalAmt"]
ws_det.append(detail_headers)
for col_num, header in enumerate(detail_headers, 1):
    cell = ws_det.cell(row=3, column=col_num)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border

for row_idx_det, row in enumerate(trans_df.itertuples(index=False), start=4):
    ws_det.append(list(row))
    for col_num in range(1, 8):
        cell = ws_det.cell(row=row_idx_det, column=col_num)
        cell.border = thin_border

output_file = "Inventory_Balance_Demo.xlsx"
wb.save(output_file)
print(f"✅ 演示报表已导出：{output_file}")
print("\n🎉 演示完成！请检查 Excel 输出格式是否符合预期。")
print("   确认无误后，使用 inventory_tracking.py 连接真实数据库运行。")
